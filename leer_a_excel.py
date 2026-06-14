#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
leer_a_excel.py

Lee hojas de respuesta OMR y vuelca los resultados directamente en el
libro Excel de analisis (hoja "2. Respuestas Estudiantes"), ubicando a
cada estudiante en la fila que corresponde a su N de lista (leido de los
recuadros de digitos de la hoja).

Estructura del Excel destino (hoja "2. Respuestas Estudiantes"):
  - Fila 15 = N de lista 1; en general  fila = 14 + N de lista.
  - Columna A = N, B = RUT (manual), C = Nombre (precargado).
  - Columna D en adelante = respuestas (P1 = col D, P2 = col E, ...).
  - NO se tocan B, C ni las columnas de formulas (BB en adelante).

Lectura del N de lista en la hoja:
  - Dos bloques de recuadros en la banda superior.
  - Bloque izquierdo: digitos 0-4 (decenas).
  - Bloque derecho: digitos 0-9 (unidades).
  - N = decena*10 + unidad.

Uso:
  python leer_a_excel.py foto.jpg --excel base.xlsx
  python leer_a_excel.py fotos/ --batch --excel base.xlsx
  python leer_a_excel.py foto.jpg --excel base.xlsx --debug

Dependencias:
  pip install opencv-python numpy openpyxl
"""

import os
import sys
import glob
import shutil
import argparse

import cv2
import numpy as np

# Reutilizamos toda la logica de vision del lector original.
import leer_omr as L

# ----------------------------------------------------------------------------
# CONFIGURACION DEL VOLCADO AL EXCEL
# ----------------------------------------------------------------------------

SHEET_NAME = "2. Respuestas Estudiantes"
FIRST_DATA_ROW = 15      # fila del N de lista 1
ROW_FOR_N = lambda n: 14 + n   # fila = 14 + N de lista
FIRST_ANSWER_COL = 4     # columna D (P1)
MAX_LIST_NUMBER = 49     # 0-4 decenas, 0-9 unidades

EXTENSIONS = L.EXTENSIONS


# ----------------------------------------------------------------------------
# LECTURA DEL N DE LISTA
# ----------------------------------------------------------------------------

def read_list_number(warped, bin_inv, debug_draw=None):
    """
    Lee el N de lista de los recuadros de digitos en la banda superior.
    Devuelve (numero:int | None, info:str).

    Estrategia:
      - Detecta las casillas en la banda superior de la hoja.
      - Las agrupa en una sola fila (la del N de lista) por su 'y'.
      - Separa en dos bloques (izq=decenas 0-4, der=unidades 0-9) por el
        gran espacio horizontal entre ambos.
      - En cada bloque, la casilla con mayor relleno (sobre umbral) es el digito.
    """
    boxes = L.detect_boxes(warped)
    H = warped.shape[0]

    # Casillas candidatas: en el tercio superior, tamano de casilla pequena.
    band = [b for b in boxes if b[1] < H * 0.22]
    if not band:
        return None, "no se detectaron recuadros en la banda superior"

    # La fila del N de lista es la banda 'y' mas poblada cerca del tope.
    ys = sorted(b[1] for b in band)
    # Agrupar por y con tolerancia.
    y_ref = ys[len(ys) // 2]
    row = [b for b in band if abs(b[1] - y_ref) <= 18]
    row.sort(key=lambda b: b[0])

    # Deben ser ~15 (5 + 10). Si hay de mas, quedarnos con los de tamano modal.
    if len(row) < 10:
        return None, f"solo {len(row)} recuadros en la fila del N de lista"

    ws = sorted(b[2] for b in row)
    mw = ws[len(ws) // 2]
    row = [b for b in row if 0.6 * mw <= b[2] <= 1.6 * mw]
    row.sort(key=lambda b: b[0])

    # Separar en dos bloques por el mayor salto horizontal entre centros.
    centers = [b[0] + b[2] / 2 for b in row]
    gaps = [(centers[i + 1] - centers[i], i) for i in range(len(centers) - 1)]
    if not gaps:
        return None, "no se pudo separar bloques"
    _, split_i = max(gaps)
    left = row[:split_i + 1]    # decenas
    right = row[split_i + 1:]   # unidades

    def marked_digit(block):
        if not block:
            return None, []
        ratios = [L.fill_ratio(bin_inv, b) for b in block]
        idx = int(np.argmax(ratios))
        return (idx if ratios[idx] >= L.FILL_THRESHOLD else None), ratios

    tens_i, _ = marked_digit(left)
    ones_i, _ = marked_digit(right)

    # Interpretacion flexible cuando falta un bloque:
    #  - decena marcada + unidad vacia  -> el valor es la decena (ej: marca "1"
    #    en decenas pensando en el numero 1). Se asume unidad = 0... salvo que
    #    eso de un multiplo de 10 sin sentido para una lista corta.
    #  - unidad marcada + decena vacia  -> unidad sola (numeros 1-9).
    if tens_i is None and ones_i is None:
        return None, "no se marco ningun digito del N de lista"

    if ones_i is None:
        # Solo decena marcada. Si la "decena" es 1-9 y no hay unidad, lo mas
        # probable es que el alumno uso ese bloque como unico digito.
        number = tens_i
        info_extra = "(solo bloque izquierdo marcado)"
    elif tens_i is None:
        number = ones_i
        info_extra = "(solo unidades)"
    else:
        number = tens_i * 10 + ones_i
        info_extra = f"(decena={tens_i}, unidad={ones_i})"

    if debug_draw is not None:
        for b in left + right:
            cv2.rectangle(debug_draw, (b[0], b[1]),
                          (b[0] + b[2], b[1] + b[3]), (255, 0, 0), 2)
        # marcar los elegidos
        if tens_i is not None:
            b = left[tens_i]
            cv2.rectangle(debug_draw, (b[0], b[1]),
                          (b[0] + b[2], b[1] + b[3]), (0, 0, 255), 3)
        if ones_i is not None:
            b = right[ones_i]
            cv2.rectangle(debug_draw, (b[0], b[1]),
                          (b[0] + b[2], b[1] + b[3]), (0, 0, 255), 3)

    if not (1 <= number <= MAX_LIST_NUMBER):
        return None, f"numero fuera de rango: {number}"
    return number, f"N={number} {info_extra}"


# ----------------------------------------------------------------------------
# PROCESAR UNA IMAGEN -> (numero, respuestas)
# ----------------------------------------------------------------------------

def process_image(path, debug=False):
    image = cv2.imread(path)
    if image is None:
        raise ValueError(f"No se pudo abrir la imagen: {path}")
    warped, method = L.deskew(image)

    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    bin_inv = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                    cv2.THRESH_BINARY_INV, 35, 15)

    dbg = warped.copy() if debug else None

    number, info = read_list_number(warped, bin_inv, debug_draw=dbg)

    dbg_path = None
    if debug:
        base = os.path.splitext(os.path.basename(path))[0]
        dbg_path = f"debug_{base}.png"
    answers = L.read_sheet(warped, debug_path=None)

    if dbg is not None:
        # dibujar tambien lectura de respuestas reusando read_sheet con su debug
        L.read_sheet(warped, debug_path=dbg_path)

    return number, answers, method, info


# ----------------------------------------------------------------------------
# VOLCADO AL EXCEL
# ----------------------------------------------------------------------------

def dump_to_excel(records, excel_path, out_path):
    """
    records: lista de (numero, answers, origen).
    Escribe en una COPIA del excel; no toca B, C ni formulas.
    """
    from openpyxl import load_workbook

    shutil.copy2(excel_path, out_path)
    wb = load_workbook(out_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{SHEET_NAME}' en {excel_path}")
    ws = wb[SHEET_NAME]

    written, skipped = [], []
    for number, answers, origen in records:
        if number is None:
            skipped.append((origen, "no se leyo N de lista"))
            continue
        row = ROW_FOR_N(number)
        for i, a in enumerate(answers):
            if a in ("DES",):           # desarrollo: dejar vacio
                continue
            val = "" if a in ("BLANCO", "AMBIGUO") else a
            ws.cell(row=row, column=FIRST_ANSWER_COL + i, value=val)
        written.append((number, row, origen))

    wb.save(out_path)
    return written, skipped


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Lector OMR -> Excel de analisis")
    ap.add_argument("entrada", help="Imagen o carpeta (con --batch)")
    ap.add_argument("--excel", required=True, help="Ruta del .xlsx de analisis")
    ap.add_argument("--out", default=None, help="Ruta del .xlsx de salida (copia)")
    ap.add_argument("--batch", action="store_true", help="Procesar carpeta")
    ap.add_argument("--debug", action="store_true", help="Imagenes de diagnostico")
    args = ap.parse_args()

    if args.batch:
        files = [f for f in glob.glob(os.path.join(args.entrada, "*"))
                 if f.lower().endswith(EXTENSIONS)]
        files.sort()
    else:
        files = [args.entrada]

    records = []
    for f in files:
        try:
            number, answers, method, info = process_image(f, debug=args.debug)
        except Exception as e:
            print(f"[ERROR] {f}: {e}")
            continue
        origen = os.path.basename(f)
        estado = info if number else f"[SIN N] {info}"
        print(f"{origen}: {estado} | {len(answers)} respuestas [{method}]")
        records.append((number, answers, origen))

    out_path = args.out or os.path.splitext(args.excel)[0] + "_RESULTADOS.xlsx"
    written, skipped = dump_to_excel(records, args.excel, out_path)

    print(f"\nVolcados {len(written)} estudiantes en: {out_path}")
    for number, row, origen in written:
        print(f"  N{number:>3} -> fila {row}  ({origen})")
    if skipped:
        print(f"\nNO volcados ({len(skipped)}) - revisar a mano:")
        for origen, motivo in skipped:
            print(f"  {origen}: {motivo}")


if __name__ == "__main__":
    main()
