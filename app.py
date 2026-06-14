#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py - Webapp local para lectura de hojas OMR y volcado a plantilla Excel.

Capa 1: funciona en tu Mac. Flujo:
  1. El usuario sube la plantilla .xlsx vacia.
  2. Ingresa la lista de estudiantes (ordenada por N de lista).
  3. Ingresa la pauta (clave + eje + habilidad + especificacion por item).
  4. Sube las fotos de las hojas de respuesta.
  5. La app procesa y devuelve el .xlsx relleno para descargar.

Ejecutar:
  python3 app.py
  luego abrir http://localhost:5000 en el navegador.

Dependencias:
  pip install flask opencv-python numpy openpyxl
"""

import os
import io
import uuid
import shutil
import tempfile

from flask import (Flask, request, render_template, send_file,
                   jsonify, session)

import cv2
import numpy as np
from openpyxl import load_workbook

import leer_omr as L
import leer_a_excel as LE

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-key-cambiar")

# Carpeta temporal por sesion de trabajo.
WORK_DIR = os.path.join(tempfile.gettempdir(), "omr_webapp")
os.makedirs(WORK_DIR, exist_ok=True)

ALLOWED_IMG = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")


def session_dir():
    """Carpeta de trabajo aislada por sesion del navegador."""
    sid = session.get("sid")
    if not sid:
        sid = uuid.uuid4().hex
        session["sid"] = sid
    d = os.path.join(WORK_DIR, sid)
    os.makedirs(d, exist_ok=True)
    return d


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/procesar", methods=["POST"])
def procesar():
    """
    Recibe: lista de estudiantes (json), pauta (json), fotos.
    Usa la plantilla incluida en la app. Devuelve resumen + id de descarga.
    """
    d = session_dir()

    # --- 1. Plantilla Excel: incluida en la app ---
    plantilla_path = os.path.join(os.path.dirname(__file__),
                                  "plantilla", "plantilla_base.xlsx")
    if not os.path.exists(plantilla_path):
        return jsonify({"error": "No se encuentra la plantilla base"}), 500

    # --- 2. Lista de estudiantes y pauta (vienen como JSON en el form) ---
    import json
    estudiantes = json.loads(request.form.get("estudiantes", "[]"))
    pauta = json.loads(request.form.get("pauta", "[]"))

    # --- 3. Fotos ---
    fotos = request.files.getlist("fotos")
    if not fotos:
        return jsonify({"error": "No se subieron fotos"}), 400

    # Procesar cada foto: leer N de lista + respuestas.
    resultados = []
    for foto in fotos:
        if not foto.filename.lower().endswith(ALLOWED_IMG):
            continue
        fpath = os.path.join(d, foto.filename)
        foto.save(fpath)
        try:
            number, answers, method, info = LE.process_image(fpath, debug=False)
        except Exception as e:
            resultados.append({"archivo": foto.filename, "n": None,
                               "estado": f"error: {e}", "respuestas": []})
            continue
        resultados.append({
            "archivo": foto.filename,
            "n": number,
            "estado": info if number else f"sin N de lista: {info}",
            "respuestas": answers,
        })

    # --- 4. Rellenar la plantilla ---
    out_path = os.path.join(d, "resultados.xlsx")
    try:
        rellenar_plantilla(plantilla_path, out_path, estudiantes, pauta,
                           resultados)
    except Exception as e:
        return jsonify({"error": f"Error al rellenar Excel: {e}"}), 500

    return jsonify({
        "ok": True,
        "resultados": [{"archivo": r["archivo"], "n": r["n"],
                        "estado": r["estado"]} for r in resultados],
        "descarga": "/descargar",
    })


def rellenar_plantilla(plantilla_path, out_path, estudiantes, pauta, resultados):
    """
    Escribe en la plantilla:
      - Nombres de estudiantes en hoja '2. Respuestas Estudiantes' col C.
      - Pauta (clave/eje/habilidad/especificacion) en hoja
        '1. Tabla de Especificaciones' cols B-E.
      - Respuestas leidas en hoja '2' desde col D, en la fila del N de lista.
    No toca formulas.
    """
    wb = load_workbook(plantilla_path)

    # --- Pauta -> Tabla de Especificaciones ---
    if "1. Tabla de Especificaciones" in wb.sheetnames and pauta:
        te = wb["1. Tabla de Especificaciones"]
        for i, item in enumerate(pauta):
            r = 16 + i  # primera fila de item
            if item.get("clave"):
                te.cell(r, 2, item["clave"])            # B
            if item.get("eje"):
                te.cell(r, 3, item["eje"])              # C
            if item.get("habilidad"):
                te.cell(r, 4, item["habilidad"])        # D
            if item.get("especificacion"):
                te.cell(r, 5, item["especificacion"])   # E

    # --- Nombres -> Respuestas Estudiantes ---
    re_sheet = wb[LE.SHEET_NAME]
    for i, nombre in enumerate(estudiantes):
        r = LE.FIRST_DATA_ROW + i
        if nombre:
            re_sheet.cell(r, 3, nombre)   # col C

    # --- Respuestas leidas ---
    for res in resultados:
        n = res["n"]
        if not n:
            continue
        r = LE.ROW_FOR_N(n)
        for j, a in enumerate(res["respuestas"]):
            if a == "DES":
                continue
            val = "" if a in ("BLANCO", "AMBIGUO") else a
            re_sheet.cell(r, LE.FIRST_ANSWER_COL + j, val)

    wb.save(out_path)


@app.route("/descargar")
def descargar():
    d = session_dir()
    out_path = os.path.join(d, "resultados.xlsx")
    if not os.path.exists(out_path):
        return "No hay archivo para descargar", 404
    return send_file(out_path, as_attachment=True,
                     download_name="resultados.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
